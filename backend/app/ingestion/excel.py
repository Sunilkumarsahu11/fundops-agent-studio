from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .models import SourceColumn, SourceLocation, SourceTable
from .type_inference import infer_type


def read_excel(content: bytes, file_name: str) -> list[SourceTable]:
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    tables: list[SourceTable] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_index = _detect_header_row(rows)
        headers = _unique_headers(rows[header_index])
        data_rows = []
        for row in rows[header_index + 1 :]:
            values = list(row[: len(headers)])
            if all(value in (None, "") for value in values):
                continue
            data_rows.append({headers[i]: values[i] for i in range(len(headers))})
        columns = []
        for index, name in enumerate(headers, start=1):
            values = [r.get(name) for r in data_rows]
            columns.append(SourceColumn(name=name, sample_values=values[:10], inferred_type=infer_type(values), nullable=any(v in (None, "") for v in values), source=SourceLocation(file_name=file_name, sheet=sheet.title, cell=f"{_column_letter(index)}{header_index + 1}")))
        tables.append(SourceTable(name=sheet.title, columns=columns, rows=data_rows, source=SourceLocation(file_name=file_name, sheet=sheet.title)))
    return tables


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    best = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        non_empty = [v for v in row if v not in (None, "")]
        text = sum(isinstance(v, str) for v in non_empty)
        score = text * 2 + len(non_empty)
        if score > best_score:
            best, best_score = index, score
    return best


def _unique_headers(row: tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for index, value in enumerate(row, start=1):
        base = str(value).strip() if value not in (None, "") else f"column_{index}"
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    while result and result[-1].startswith("column_") and not any(v not in (None, "") for v in row[len(result):]):
        break
    return result


def _column_letter(number: int) -> str:
    value = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        value = chr(65 + remainder) + value
    return value
