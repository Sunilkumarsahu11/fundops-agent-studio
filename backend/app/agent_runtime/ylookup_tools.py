from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

MONEY_TOLERANCE = Decimal("0.01")


@dataclass(frozen=True)
class BankTransaction:
    account_name: str
    account_number: str
    currency: str
    bank_reference: str
    customer_reference: str
    trn_type: str
    value_date: str
    credit: Decimal
    debit: Decimal
    balance: Decimal
    narrative: str

    @property
    def signed_amount(self) -> Decimal:
        return self.credit + self.debit


def _dec(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _headers(ws, header_row: int = 1) -> list[str]:
    return [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]


def _rows(ws, header_row: int = 1):
    headers = _headers(ws, header_row)
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None for v in values):
            continue
        yield dict(zip(headers, values))


def _dedupe_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    import json
    for issue in issues:
        key = json.dumps(issue, sort_keys=True, default=str)
        if key not in seen:
            seen.add(key)
            result.append(issue)
    return result


def workbook_sheet_summary(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result = []
    for ws in wb.worksheets:
        headers = _headers(ws) if ws.max_row else []
        nonempty = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                nonempty += 1
        result.append({"sheet": ws.title, "columns": headers, "row_count": nonempty})
    return result


def _parse_bank_pdf(content: bytes) -> tuple[dict[str, Any], list[BankTransaction]]:
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("PyMuPDF is required for bank statement PDF parsing") from exc

    doc = fitz.open(stream=content, filetype="pdf")
    text = "\n".join(page.get_text("text") for page in doc)
    meta: dict[str, Any] = {}
    for key, pattern in {
        "account_name": r"Account name\s+(.+)",
        "account_number": r"Account number\s+([\w-]+)",
        "currency": r"Currency\s+([A-Z]{3})",
        "iban": r"IBAN\s+([A-Z0-9]+)",
        "bank_name": r"Bank name\s+(.+)",
    }.items():
        m = re.search(pattern, text)
        if m:
            meta[key] = m.group(1).strip()

    txns: list[BankTransaction] = []
    lines = [x.rstrip() for x in text.splitlines()]
    for i in range(0, len(lines) - 7):
        ref, customer, trn, date = [lines[i + k].strip() for k in range(4)]
        if not re.fullmatch(r"\d{2} \w{3} \d{4}", date):
            continue
        if not re.fullmatch(r"(?:TT\s+)?[A-Z0-9_-]+", ref) or not customer or not trn:
            continue
        amount = _dec(lines[i + 4])
        balance = _dec(lines[i + 5])
        if amount is None or balance is None or not re.fullmatch(r"\d{2}:\d{2}", lines[i + 6].strip()):
            continue
        post_date = lines[i + 7].strip()
        if not re.fullmatch(r"\d{2} \w{3} \d{4}", post_date):
            continue
        narrative = ""
        if i + 9 < len(lines) and lines[i + 8].strip() == "Narrative":
            narrative = lines[i + 9].strip()
        credit = amount if amount >= 0 else Decimal(0)
        debit = amount if amount < 0 else Decimal(0)
        txns.append(BankTransaction(meta.get("account_name", ""), meta.get("account_number", ""), meta.get("currency", ""), ref, customer, trn, date, credit, debit, balance, narrative))
    return meta, txns


def bank_statement_review_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Parse a Ylookup bank statement and run deterministic statement controls."""
    import base64
    file_name = str(inputs.get("file_name", ""))
    raw = base64.b64decode(inputs["content_base64"], validate=True)
    if not file_name.lower().endswith(".pdf"):
        raise ValueError("bank_statement_review requires a PDF bank statement")
    meta, txns = _parse_bank_pdf(raw)
    exceptions: list[dict[str, Any]] = []
    seen_refs: set[tuple[str, str]] = set()
    for idx, t in enumerate(txns, start=1):
        if t.credit != 0 and t.debit != 0:
            exceptions.append({"code": "BOTH_CREDIT_AND_DEBIT", "row": idx})
        key = (t.bank_reference, t.value_date)
        if key in seen_refs and t.bank_reference != "NONREF":
            exceptions.append({"code": "DUPLICATE_BANK_REFERENCE", "row": idx, "bank_reference": t.bank_reference})
        seen_refs.add(key)
    balance_breaks = 0
    for current, previous in zip(txns, txns[1:]):
        expected = previous.balance + current.signed_amount
        if abs(expected - current.balance) > MONEY_TOLERANCE:
            balance_breaks += 1
            exceptions.append({"code": "BALANCE_ROLLFORWARD_BREAK", "row_balance": str(current.balance), "expected": str(expected), "reference": current.bank_reference})
    unique_exceptions = _dedupe_issues(exceptions)
    return {
        "status": "exceptions" if unique_exceptions else "passed",
        "file_name": file_name,
        "account": meta,
        "transaction_count": len(txns),
        "balance_break_count": balance_breaks,
        "transactions": [
            {"bank_reference": t.bank_reference, "customer_reference": t.customer_reference, "trn_type": t.trn_type, "value_date": t.value_date, "credit": str(t.credit), "debit": str(t.debit), "balance": str(t.balance), "narrative": t.narrative}
            for t in txns
        ],
        "exceptions": unique_exceptions,
    }


def bank_workbook_control_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Validate the Ylookup bank-statement working workbook against reference sheets."""
    import base64
    raw = base64.b64decode(inputs["content_base64"], validate=True)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    staging = wb["Staging Sheet"]
    account_map = {str(r.get("Account Number")): str(r.get("Bank Account")) for r in _rows(wb["Account Map"]) if r.get("Account Number")}
    legal_entities = {str(r.get("Legal Entity")) for r in _rows(wb["Legal Entity Master List"]) if r.get("Legal Entity")}
    vendors = {str(r.get("Vendor")) for r in _rows(wb["Vendor Master List"]) if r.get("Vendor")}
    project_codes = {str(r.get("Project Code")) for r in _rows(wb["Project Code Report"]) if r.get("Project Code")}
    related = {str(r.get("Related Party")) for r in _rows(wb["Related Party Master"]) if r.get("Related Party")}
    exceptions: list[dict[str, Any]] = []
    count = 0
    for row in _rows(staging):
        count += 1
        account = str(row.get("Account Number") or "")
        entity = str(row.get("Matched Legal Entity") or "")
        if account and account not in account_map:
            exceptions.append({"code": "UNKNOWN_BANK_ACCOUNT", "account_number": account})
        if entity and entity not in legal_entities:
            exceptions.append({"code": "UNKNOWN_LEGAL_ENTITY", "entity": entity})
        matched_vendor = row.get("Matched Sender/Beneficiary")
        if matched_vendor and str(matched_vendor) not in vendors and str(matched_vendor) not in legal_entities:
            exceptions.append({"code": "UNKNOWN_COUNTERPARTY", "counterparty": str(matched_vendor)})
        project = row.get("Matched Project Code")
        if project and str(project) not in project_codes:
            exceptions.append({"code": "UNKNOWN_PROJECT_CODE", "project_code": str(project)})
        related_party = row.get("Related Party Match")
        if related_party and str(related_party) not in related:
            exceptions.append({"code": "UNKNOWN_RELATED_PARTY", "related_party": str(related_party)})
        if row.get("Credit amount") not in (None, "") and row.get("Debit amount") not in (None, ""):
            exceptions.append({"code": "BOTH_CREDIT_AND_DEBIT", "account_number": account})
    unique_exceptions = _dedupe_issues(exceptions)
    return {"status": "exceptions" if unique_exceptions else "passed", "row_count": count, "raw_exception_count": len(exceptions), "exception_count": len(unique_exceptions), "exceptions": unique_exceptions}


def journal_entry_control_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Check Ylookup DIU journal-entry pairs for double-entry balance and source linkage."""
    import base64
    raw = base64.b64decode(inputs["content_base64"], validate=True)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = list(_rows(wb["DIU "]))
    groups: dict[tuple[Any, Any], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        groups[(r.get("Batch ID"), r.get("JE Index"))].append(r)
    exceptions = []
    for key, batch in groups.items():
        debits = sum((_dec(r.get("Amount (Local)")) or Decimal(0)) for r in batch if str(r.get("is Debit", "")).lower() == "yes")
        credits = sum((_dec(r.get("Amount (Local)")) or Decimal(0)) for r in batch if str(r.get("is Debit", "")).lower() == "no")
        if abs(debits - credits) > MONEY_TOLERANCE:
            exceptions.append({"code": "UNBALANCED_JOURNAL_ENTRY", "batch": key, "debits": str(debits), "credits": str(credits), "difference": str(debits - credits)})
        if len(batch) < 2:
            exceptions.append({"code": "INCOMPLETE_JOURNAL_ENTRY", "batch": key, "line_count": len(batch)})
        refs = {str(r.get("Transaction Reference") or "") for r in batch}
        if len(refs) != 1:
            exceptions.append({"code": "JOURNAL_REFERENCE_MISMATCH", "batch": key, "references": sorted(refs)})
    return {"status": "exceptions" if exceptions else "passed", "journal_entry_count": len(groups), "exception_count": len(exceptions), "exceptions": exceptions}


def investor_loader_control_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Validate investor-level GL rows against the supplied loader/reference workbook."""
    import base64
    raw_gl = base64.b64decode(inputs["gl_content_base64"], validate=True)
    raw_ref = base64.b64decode(inputs["reference_content_base64"], validate=True)
    gl_wb = load_workbook(io.BytesIO(raw_gl), read_only=True, data_only=True)
    ref_wb = load_workbook(io.BytesIO(raw_ref), read_only=True, data_only=True)
    gl_rows = list(_rows(gl_wb["Investor-Level GL"]))
    le_map = {(str(r.get("Legal Entity")), str(r.get("Currency"))): str(r.get("Corvus LE")) for r in _rows(ref_wb["LE Mapping"], header_row=2) if r.get("Legal Entity")}
    inv_map = {str(r.get("Investor Lookup")): str(r.get("Ext Ref Check")) for r in _rows(ref_wb["Investor Mapping"]) if r.get("Investor Lookup")}
    deal_map = {(str(r.get("Deal Name")), str(r.get("Currency"))): str(r.get("Curr Check")) for r in _rows(ref_wb["Deal Mapping"]) if r.get("Deal Name")}
    coa_map = {(str(r.get("Helio GL Account")), str(r.get("Helio Trans Type"))): str(r.get("Verado II GL Account Code")) for r in _rows(ref_wb["CoA Mapping"]) if r.get("Helio GL Account") or r.get("Helio Trans Type")}
    exceptions: list[dict[str, Any]] = []
    batch_groups: dict[tuple[Any, Any, Any], list[dict[str, Any]]] = defaultdict(list)
    for row in gl_rows:
        entity, currency = str(row.get("Legal Entity") or ""), str(row.get("Transaction Currency") or "")
        if entity and (entity, currency) not in le_map:
            exceptions.append({"code": "LEGAL_ENTITY_CURRENCY_MAPPING_GAP", "entity": entity, "currency": currency})
        inv = row.get("Investor")
        if inv:
            lookup = f"{entity}.{inv}"
            if lookup not in inv_map:
                exceptions.append({"code": "INVESTOR_MAPPING_GAP", "investor_lookup": lookup})
        deal = row.get("Deal Name")
        if deal and (str(deal), currency) not in deal_map:
            exceptions.append({"code": "DEAL_CURRENCY_MAPPING_GAP", "deal": str(deal), "currency": currency})
        gl = str(row.get("GL Account") or "")
        tt = str(row.get("Trans Type") or "")
        if (gl, tt) not in coa_map:
            exceptions.append({"code": "COA_MAPPING_GAP", "gl_account": gl, "trans_type": tt})
        batch_groups[(row.get("Batch ID"), row.get("Journal Entry Index"), entity)].append(row)
    unbalanced = 0
    for key, batch in batch_groups.items():
        debit = sum((_dec(r.get("Debits (Entity Currency)")) or Decimal(0)) for r in batch)
        credit = sum((_dec(r.get("Credits (Entity Currency)")) or Decimal(0)) for r in batch)
        if abs(debit - credit) > MONEY_TOLERANCE:
            unbalanced += 1
            exceptions.append({"code": "UNBALANCED_INVESTOR_JE", "batch": key, "debits": str(debit), "credits": str(credit)})
    unique_exceptions = _dedupe_issues(exceptions)
    return {"status": "exceptions" if unique_exceptions else "passed", "row_count": len(gl_rows), "journal_entry_count": len(batch_groups), "unbalanced_journal_entries": unbalanced, "raw_exception_count": len(exceptions), "exception_count": len(unique_exceptions), "exceptions": unique_exceptions}


def mapping_gap_control_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Return deterministic mapping gaps from the verified loader workbook."""
    import base64
    raw = base64.b64decode(inputs["content_base64"], validate=True)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    gaps = []
    if "Mapping Gaps" in wb.sheetnames:
        for r in _rows(wb["Mapping Gaps"]):
            if r.get("GL Account") or r.get("Trans Type"):
                gaps.append({"gl_account": r.get("GL Account"), "trans_type": r.get("Trans Type"), "row_count": r.get("Row Count"), "total_amount": r.get("Total Amount (Entity Currency)"), "proposed_account": r.get("Proposed Verado II Account"), "proposed_trans_type": r.get("Proposed Verado II TransType"), "approval": r.get("Approval")})
    return {"status": "exceptions" if gaps else "passed", "mapping_gap_count": len(gaps), "gaps": gaps}


def movements_control_tool(inputs: dict[str, Any]) -> dict[str, Any]:
    """Check the verified loader movement reconciliation."""
    import base64
    raw = base64.b64decode(inputs["content_base64"], validate=True)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if "Movements Rec" not in wb.sheetnames:
        raise ValueError("Movements Rec sheet not found")
    exceptions = []
    rows = list(_rows(wb["Movements Rec"]))
    for r in rows:
        debits = _dec(r.get("Sum Debits")) or Decimal(0)
        credits = _dec(r.get("Sum Credits")) or Decimal(0)
        net = _dec(r.get("Net Movement")) or Decimal(0)
        if abs((debits - credits) - net) > MONEY_TOLERANCE:
            exceptions.append({"code": "MOVEMENT_RECONCILIATION_BREAK", "legal_entity": r.get("Legal Entity"), "account": r.get("Verado II GL Account"), "debits": str(debits), "credits": str(credits), "net_movement": str(net)})
    return {"status": "exceptions" if exceptions else "passed", "row_count": len(rows), "exception_count": len(exceptions), "exceptions": exceptions}
