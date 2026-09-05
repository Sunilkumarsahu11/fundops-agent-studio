from __future__ import annotations
import base64, io
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any
from openpyxl import load_workbook
MONEY_TOLERANCE=Decimal('0.01')

def _dec(v:Any)->Decimal:
    if v in (None,''): return Decimal('0')
    try: return Decimal(str(v).replace(',','').strip())
    except (InvalidOperation,ValueError) as e: raise ValueError(f'Invalid monetary value: {v!r}') from e

def _b64(v:str)->bytes:
    try: return base64.b64decode(v,validate=True)
    except Exception as e: raise ValueError('content_base64 is not valid base64') from e

def _headers(ws,row=1): return [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=row,max_row=row,values_only=True))]
def _rows(ws,row=1):
    h=_headers(ws,row)
    for vals in ws.iter_rows(min_row=row+1,values_only=True):
        if any(v is not None for v in vals): yield dict(zip(h,vals))
def _clean(v): return v.isoformat() if hasattr(v,'isoformat') else v

def bank_pdf_to_canonical_tool(inputs:dict[str,Any])->dict[str,Any]:
    from .ylookup_tools import _parse_bank_pdf
    name=str(inputs.get('file_name',''))
    if not name.lower().endswith('.pdf'): raise ValueError('bank_pdf_to_canonical requires a PDF')
    meta,txns=_parse_bank_pdf(_b64(inputs['content_base64']))
    records=[]
    for i,t in enumerate(txns,1):
        records.append({'record_id':f'{name}:{i}','record_type':'bank_transaction','account_name':t.account_name,'account_number':t.account_number,'currency':t.currency,'bank_reference':t.bank_reference,'customer_reference':t.customer_reference,'transaction_type':t.trn_type,'value_date':t.value_date,'credit':str(t.credit),'debit':str(t.debit),'amount':str(t.signed_amount),'balance':str(t.balance),'narrative':t.narrative,'source':{'file':name,'record_index':i}})
    return {'source':meta|{'file_name':name},'record_count':len(records),'records':records}

def bank_transactions_to_journal_entries_tool(inputs:dict[str,Any])->dict[str,Any]:
    wb=load_workbook(io.BytesIO(_b64(inputs['workbook_content_base64'])),read_only=True,data_only=True)
    staging=list(_rows(wb['Staging Sheet'])); coa=list(_rows(wb['CoA'])); allocations=list(_rows(wb['Allocation Rule']))
    account_map={str(r.get('Account Number')):r.get('Bank Account') for r in _rows(wb['Account Map']) if r.get('Account Number')}
    by_ref=defaultdict(list)
    for r in staging: by_ref[str(r.get('Bank reference') or '')].append(r)
    alloc={str(r.get('Legal Entity')):r.get('Allocation Rule') for r in allocations if r.get('Legal Entity')}
    journals=[]; exceptions=[]
    for record in inputs.get('records',[]):
        ref=str(record.get('bank_reference') or ''); candidates=by_ref.get(ref,[])
        if not candidates: exceptions.append({'code':'NO_STAGING_MATCH','record_id':record.get('record_id'),'bank_reference':ref}); continue
        if len(candidates)>1 and ref!='NONREF': exceptions.append({'code':'AMBIGUOUS_STAGING_MATCH','record_id':record.get('record_id'),'bank_reference':ref,'candidate_count':len(candidates)}); continue
        s=candidates[0]; entity=s.get('Matched Legal Entity') or s.get('Account Name'); cash=s.get('Cash Leg Transtype'); counter=s.get('Counterparty Transtype')
        if not cash or not counter: exceptions.append({'code':'TRANSACTION_TYPE_UNRESOLVED','record_id':record.get('record_id'),'cash_leg':cash,'counterparty_leg':counter}); continue
        amount=abs(_dec(record.get('amount'))); debit=_dec(record.get('debit'))!=0
        common={'legal_entity':entity,'currency':s.get('Currency'),'deal_name':s.get('Resolved Deal'),'position':s.get('Resolved Position'),'gl_date':_clean(s.get('Value date ')),'effective_date':_clean(s.get('Value date ')),'allocation_rule':alloc.get(str(entity)),'project_code':s.get('Matched Project Code'),'vendor':s.get('Matched Sender/Beneficiary'),'related_party':s.get('Related Party Match'),'bank_account':account_map.get(str(s.get('Account Number'))),'transaction_reference':ref,'comments':s.get('Narrative'),'source':{'bank_record_id':record.get('record_id'),'staging_reference':ref}}
        journals += [common|{'line':1,'leg':'cash','is_debit':'Y' if debit else 'N','amount_local':str(amount),'trans_type':cash},common|{'line':2,'leg':'counterparty','is_debit':'N' if debit else 'Y','amount_local':str(amount),'trans_type':counter}]
    known={str(r.get('Trans Type')) for r in coa if r.get('Trans Type')}
    for j in journals:
        if j['trans_type'] not in known: exceptions.append({'code':'COA_TRANSACTION_TYPE_GAP','transaction_type':j['trans_type'],'record_id':j['source']['bank_record_id']})
    return {'status':'exceptions' if exceptions else 'passed','record_count':len(inputs.get('records',[])),'journal_line_count':len(journals),'journal_batch_count':len({j['transaction_reference'] for j in journals}),'journals':journals,'exceptions':exceptions}

def investor_gl_to_loader_tool(inputs:dict[str,Any])->dict[str,Any]:
    gl=load_workbook(io.BytesIO(_b64(inputs['gl_content_base64'])),read_only=True,data_only=True); ref=load_workbook(io.BytesIO(_b64(inputs['reference_content_base64'])),read_only=True,data_only=True)
    gl_rows=list(_rows(gl['Investor-Level GL'])); le_rows=list(_rows(ref['LE Mapping'],2)); inv_rows=list(_rows(ref['Investor Mapping'])); deal_rows=list(_rows(ref['Deal Mapping'])); coa_rows=list(_rows(ref['CoA Mapping'])); batch_rows=list(_rows(ref['Batch Preference']))
    le={(str(r.get('Legal Entity')),str(r.get('Currency'))):r for r in le_rows}; inv={(str(r.get('Legal Entity')),str(r.get('Investor'))):r for r in inv_rows if r.get('Legal Entity') and r.get('Investor')}; deals={(str(r.get('Deal Name')),str(r.get('Position') or ''),str(r.get('Currency'))):r for r in deal_rows}; coa={(str(r.get('Helio GL Account')),str(r.get('Helio Trans Type'))):r for r in coa_rows}; priority={str(r.get('Batch Type')):r.get('Prioritization') for r in batch_rows}
    rows=[]; unresolved=[]; exceptions=[]
    for n,r in enumerate(gl_rows,1):
        entity=str(r.get('Legal Entity') or ''); cur=str(r.get('Transaction Currency') or ''); investor=str(r.get('Investor') or ''); deal=str(r.get('Deal Name') or ''); pos=str(r.get('Position') or ''); gla=str(r.get('GL Account') or ''); tt=str(r.get('Trans Type') or '')
        lm=le.get((entity,cur)); im=inv.get((entity,investor)) if investor else None; dm=deals.get((deal,pos,cur)) or deals.get((deal,'',cur)); cm=coa.get((gla,tt)); missing=[]
        if lm is None: missing.append('legal_entity_currency')
        if investor and im is None: missing.append('investor')
        if deal and dm is None: missing.append('deal_position_currency')
        if cm is None: missing.append('coa')
        if missing: unresolved.append({'source_row':n,'batch_id':r.get('Batch ID'),'transaction_index':r.get('Transaction Index'),'missing':missing,'legal_entity':entity,'investor':investor,'deal':deal,'position':pos,'gl_account':gla,'trans_type':tt}); continue
        debit=_dec(r.get('Debits (Local Currency)')); is_debit='Y' if debit!=0 else 'N'; mapped=cm.get('Verado II TransType (Debit)') if is_debit=='Y' and cm.get('Verado II TransType (Debit)') else cm.get('Verado II TransType (Default)')
        if not mapped: exceptions.append({'code':'MAPPED_TRANSACTION_TYPE_EMPTY','source_row':n,'gl_account':gla,'trans_type':tt}); continue
        rows.append({'batch_index':r.get('Batch ID'),'je_index':r.get('Journal Entry Index'),'transaction_index':r.get('Transaction Index'),'legal_entity':lm.get('Corvus LE'),'legal_entity_id':lm.get('Corvus LE ID'),'gl_date':_clean(r.get('GL Date')),'effective_date':_clean(r.get('Effective Date')),'deal_name':dm.get('New Deal Name'),'deal_id':dm.get('Corvus Deal ID'),'position':dm.get('New Position Name'),'position_id':dm.get('Corvus Position ID'),'trans_type':mapped,'transaction_currency':cur,'investor_amount_local':str(abs(_dec(r.get('Amount (Local Currency)')))),'is_debit':is_debit,'investor_amount_le':str(abs(_dec(r.get('Amount (Entity Currency)')))),'batch_type':r.get('Batch Type'),'batch_priority':priority.get(str(r.get('Batch Type'))),'batch_comments':r.get('Comments Batch'),'transaction_comments':r.get('Comments transaction'),'allocation_rule':r.get('Allocation Rule'),'investor_account_id':im.get('Corvus Specific Id') if im else None,'vehicle':im.get('Corvus Veh Name') if im else r.get('Vehicle'),'bank_account':r.get('Bank Account'),'udf_lookup':cm.get('UDF Lookup'),'supplier':r.get('Supplier Tag'),'investor_quantity':r.get('Quantity'),'source':{'file':inputs.get('gl_file_name'),'source_row':n,'batch_id':r.get('Batch ID'),'transaction_index':r.get('Transaction Index')}}
    return {'status':'exceptions' if unresolved or exceptions else 'passed','source_row_count':len(gl_rows),'loader_row_count':len(rows),'rows':rows,'mapping_exceptions':exceptions,'unresolved_rows':unresolved}

def loader_rows_from_workbook_tool(inputs:dict[str,Any])->dict[str,Any]:
    wb=load_workbook(io.BytesIO(_b64(inputs['content_base64'])),read_only=True,data_only=True); sheet=inputs.get('sheet_name','Upload Template (VERIFIED v4c)'); ws=wb[sheet]; rows=[{k:_clean(v) for k,v in r.items() if k} for r in _rows(ws)]; return {'sheet':sheet,'row_count':len(rows),'rows':rows}

def loader_reconciliation_tool(inputs:dict[str,Any])->dict[str,Any]:
    generated=inputs.get('generated_rows',[]); verified=inputs.get('verified_rows',[]); keys=inputs.get('key_fields',['batch_index','je_index','transaction_index','legal_entity','trans_type']); amount=inputs.get('amount_field','investor_amount_local')
    def key(r): return tuple(str(r.get(k) or '') for k in keys)
    left=defaultdict(list); right=defaultdict(list)
    for r in generated:left[key(r)].append(r)
    for r in verified:right[key(r)].append(r)
    ex=[]
    for k in sorted(set(left)|set(right)):
        l,r=left.get(k,[]),right.get(k,[])
        if len(l)!=len(r): ex.append({'code':'ROW_COUNT_MISMATCH','key':k,'generated':len(l),'verified':len(r)})
        for a,b in zip(l,r):
            d=_dec(a.get(amount))-_dec(b.get(amount))
            if abs(d)>MONEY_TOLERANCE: ex.append({'code':'AMOUNT_MISMATCH','key':k,'generated':str(a.get(amount)),'verified':str(b.get(amount)),'difference':str(d)})
    return {'status':'exceptions' if ex else 'passed','generated_rows':len(generated),'verified_rows':len(verified),'exception_count':len(ex),'exceptions':ex}
