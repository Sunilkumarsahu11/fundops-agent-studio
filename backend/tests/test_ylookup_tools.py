import base64
import io

from openpyxl import Workbook

from app.agent_runtime.ylookup_tools import (
    bank_workbook_control_tool,
    investor_loader_control_tool,
    journal_entry_control_tool,
    mapping_gap_control_tool,
    movements_control_tool,
)


def _b64(wb: Workbook) -> str:
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def test_journal_entry_control_passes_balanced_pair() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DIU "
    ws.append(["Batch ID", "JE Index", "is Debit", "Amount (Local)", "Transaction Reference"])
    ws.append([1, 1, "Yes", 100, "REF1"])
    ws.append([1, 1, "No", 100, "REF1"])
    result = journal_entry_control_tool({"content_base64": _b64(wb)})
    assert result["status"] == "passed"
    assert result["journal_entry_count"] == 1


def test_mapping_gap_control_reads_unresolved_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Gaps"
    ws.append(["GL Account", "Trans Type", "Row Count", "Total Amount (Entity Currency)", "Proposed Verado II Account", "Proposed Verado II TransType", "Approval"])
    ws.append(["40070", "Expense: Administration Fees", 11, 4867.16, "50080", "Expense: Admin fees", None])
    result = mapping_gap_control_tool({"content_base64": _b64(wb)})
    assert result["status"] == "exceptions"
    assert result["mapping_gap_count"] == 1


def test_movements_control_passes_when_net_equals_debits_less_credits() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movements Rec"
    ws.append(["Legal Entity", "Verado II GL Account", "Sum Debits", "Sum Credits", "Net Movement"])
    ws.append(["Fund A", "10000 - Cash", 150, 40, 110])
    result = movements_control_tool({"content_base64": _b64(wb)})
    assert result["status"] == "passed"


def test_bank_workbook_control_detects_unknown_reference() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Staging Sheet"
    ws.append(["Account Number", "Matched Legal Entity", "Matched Sender/Beneficiary", "Matched Project Code", "Related Party Match", "Credit amount", "Debit amount"])
    ws.append(["999", "Unknown Fund", "Unknown Vendor", "BAD", "BAD-RP", 10, None])
    for title, header, value in [
        ("Account Map", ["Account Number", "Bank Account"], ["100", "Bank"]),
        ("Legal Entity Master List", ["Legal Entity"], ["Fund A"]),
        ("Vendor Master List", ["Legal Entity Domain", "Vendor"], ["NIP", "Vendor A"]),
        ("Project Code Report", ["Legal Entity Domain", "Project Code", "New Project Code"], ["NIP", "GOOD", None]),
        ("Related Party Master", ["Legal Entity Domain", "Related Party"], ["NIP", "RP A"]),
    ]:
        w = wb.create_sheet(title); w.append(header); w.append(value)
    result = bank_workbook_control_tool({"content_base64": _b64(wb)})
    codes = {x["code"] for x in result["exceptions"]}
    assert "UNKNOWN_BANK_ACCOUNT" in codes
    assert "UNKNOWN_LEGAL_ENTITY" in codes


def test_investor_loader_control_balances_minimal_gl() -> None:
    gl = Workbook(); ws = gl.active; ws.title = "Investor-Level GL"
    headers = ["Legal Entity", "Transaction Currency", "Investor", "Deal Name", "GL Account", "Trans Type", "Batch ID", "Journal Entry Index", "Debits (Entity Currency)", "Credits (Entity Currency)"]
    ws.append(headers); ws.append(["Fund A", "USD", None, "Deal A", "10000", "Cash", 1, 1, 100, 0]); ws.append(["Fund A", "USD", None, "Deal A", "20000", "Capital", 1, 1, 0, 100])
    ref = Workbook()
    le = ref.active; le.title = "LE Mapping"; le.append(["Fund Family", "Legal Entity", "Currency", "Corvus LE", "Corvus LE ID", "Corvus Currency"]); le.append(["X", "Fund A", "USD", "Fund A", 1, "USD"])
    inv = ref.create_sheet("Investor Mapping"); inv.append(["Investor Lookup", "Ext Ref Check"])
    deal = ref.create_sheet("Deal Mapping"); deal.append(["Deal Name", "New Deal Name", "Position", "New Position Name", "Currency", "Corvus Deal Name", "Corvus Deal ID", "Corvus Position Name", "Corvus Position ID", "Currency2", "Curr Check"]); deal.append(["Deal A", "Deal A", None, None, "USD", "Deal A", 1, None, None, "USD", "OK"])
    coa = ref.create_sheet("CoA Mapping"); coa.append(["Helio GL Account", "Helio Trans Type", "Verado II GL Account Code"]); coa.append(["10000", "Cash", "10000"]); coa.append(["20000", "Capital", "20000"])
    result = investor_loader_control_tool({"gl_content_base64": _b64(gl), "reference_content_base64": _b64(ref)})
    assert result["unbalanced_journal_entries"] == 0
